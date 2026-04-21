"""Generate cards.json for the GitHub Pages gallery from Draft Cube.xlsx."""
import json
import os
import re
import sys
import urllib.request
from pathlib import Path

import openpyxl

from _utils import norm

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "Draft Cube.xlsx"
KRCG_CACHE = ROOT / "data" / "krcg_vtes.json"
KRCG_URL = "https://static.krcg.org/data/vtes.json"
DRAFT_OCR = ROOT / "data" / "draft_ocr.json"
DRAFT_OVERRIDES = ROOT / "data" / "draft_overrides.json"
CRYPT_DIR = "images/crypt"
LIB_DIRS = {
    "Common": "images/library-common",
    "Uncommon": "images/library-uncommon",
    "Rare": "images/library-rare",
}
OUT = ROOT / "data" / "cards.json"


def index_dir(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    for f in os.listdir(path):
        stem = os.path.splitext(f)[0]
        out.setdefault(stem, f)
    return out


def find_crypt(stem_index: dict[str, str], name: str) -> str | None:
    n = norm(name)
    if n in stem_index:
        return stem_index[n]
    pat = re.compile(rf"^{re.escape(n)}g\d+$")
    for stem, fname in stem_index.items():
        if pat.match(stem):
            return fname
    return None


def load_krcg() -> list[dict]:
    if not KRCG_CACHE.exists():
        print(f"Downloading KRCG data → {KRCG_CACHE.name}")
        req = urllib.request.Request(KRCG_URL, headers={"User-Agent": "Vtes.Draft/1.0"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            KRCG_CACHE.write_bytes(resp.read())
    return json.loads(KRCG_CACHE.read_text(encoding="utf-8"))


def clean_draft_snippet(snip: str) -> str:
    """Trim OCR noise around the DRAFT clause + fix common OCR artefacts."""
    if not snip:
        return ""
    up = snip.upper()
    i = up.find("DRAFT:")
    if i < 0:
        i = up.find("DRAFT;")
    if i < 0:
        return ""
    s = snip[i:]

    # Cut at card-footer markers (illustrator / copyright)
    end_markers = [
        "Illus:", "Illus ", "Illus;", "Illu:", "Ius:", "lus:", "llus:", "ilus:",
        "©", "\u24B8", "(@", "@ 20", " 0 20", " 0 202", "02024", "02025", "0 1024", "0 2024",
        "Paradox", "Parador", "(publ", "publ)", "publl", "(puby",
    ]
    cut = len(s)
    for m in end_markers:
        pos = s.find(m)
        if pos > 0:
            cut = min(cut, pos)
    s = s[:cut]

    # Normalize DRAFT prefix
    s = re.sub(r"^DRAFT\s*[;:]", "DRAFT:", s, flags=re.I)

    # Strip @ icon variants aggressively (don't absorb real capitalized words)
    s = re.sub(r"[a-z]?@[\}\)\]]+", "", s)         # l@}, @} (must come before @ alone)
    s = re.sub(r"@(?=[A-Z][a-z])", "", s)          # @As -> As, @Reduce -> Reduce
    s = re.sub(r"@[A-Z](?=As\b)", "", s)           # @JAs -> As (icon letter before As)
    s = re.sub(r"@[A-Z]{0,2}\]?", "", s)           # @EJ, @J, @], @

    # Other icon-letter garbage tokens
    icon_tokens = [
        r"\[[A-Za-z\d]{1,3}\]?",          # [U, [d], [1]
        r"[A-Z]J(?=\s|$|[A-Z])",           # DJ, EJ at word-end
        r"[A-Z]\](?=\s|$)",                # D], E]
        r"Xis\s*\d*\.?",                   # Xis 0.
        r"\bEE\b",
        r"[€£¥§¢][A-Z]{0,2}",             # currency-symbol OCR of icon
    ]
    icon_re = "|".join(icon_tokens)

    s = re.sub(rf"^(DRAFT:)\s*(?:{icon_re})\s*", r"\1 ", s)
    s = re.sub(rf"(?<=[.\s])(?:{icon_re})\s+", " ", s)

    # Trailing mixed-case short gibberish after "above" (e.g. "above UeDl")
    s = re.sub(r"\babove\s+[A-Za-z]{1,6}\s*\.?$", "above.", s, flags=re.I)

    # Trailing illustrator name leaked as plain text (e.g. "ur Steve Prescott", "Randi Galleeos")
    s = re.sub(r"\s+\w{1,3}\s+[A-Z][a-z]+\s+[A-Z][a-z]+\s*\.?$", "", s)
    s = re.sub(r"\s+[A-Z][a-z]+\s+[A-Z][a-z]+\s*\.?$", ".", s)

    # Bracket stuck to word: [above -> above, word] -> word
    s = re.sub(r"\[(?=[A-Za-z])", "", s)
    s = re.sub(r"(?<=[A-Za-z])[\]\}]", "", s)

    # Illustrator marker without colon (Ius David Day, Ius James Stowe)
    s = re.sub(r"\s+Ius\s+[A-Z].*$", "", s)

    # Tail copyright garbage: "(,20XX ..." or "0 20XX ..." not caught by end_markers
    s = re.sub(r"\s+\(?,?\s*20\d{2}\b.*$", "", s)
    s = re.sub(r"\s+0\s+1?02[45].*$", "", s)

    # Flavor text attribution (— Name or a trailing " The <Capital word>...")
    s = re.sub(r"\s+[\u2014\u2013-]\s*[A-Z][a-z]+.*$", "", s)

    # Missing "As" before "above" at the start
    s = re.sub(r"^DRAFT:\s+above\b", "DRAFT: As above", s)

    # Icon letter(s) stuck to "above" (Mlabove, Eilabove, EJabove, labove, Jabove, ...)
    s = re.sub(r"\b(?:M[il]?|E[il]{0,2}|EJ|[ML]|la|[A-Z])above\b", "above", s, flags=re.I)

    # Icon letter(s) stuck to "As" (DAs, EAs, EYAs, JAs -> As)
    s = re.sub(r"\b[A-Z]{1,3}(?=As\b)", "", s)

    # Double-capital word starts (icon + real letter: EEnter -> Enter)
    s = re.sub(r"\b([A-Z])\1(?=[a-z])", r"\1", s)

    # Leading single-letter icon right after "DRAFT:"
    s = re.sub(r"^DRAFT:\s+[A-Z]\s+(?=[A-Za-z])", "DRAFT: ", s)
    # Leading single-letter icon inside clause after period/space
    s = re.sub(r"(?<=[.\s])[A-Z]\s+(?=[A-Z][a-z])", " ", s)

    # Trailing single-letter icon (" M.", " I.", " J.")
    s = re.sub(r"\s+[A-Z]\.?\s*$", "", s)
    s = re.sub(r"\s+[A-Z]\b(?=\s*\.?$)", "", s)

    # Pipe misread for '1' in numeric contexts
    s = re.sub(r"\+\s*[|\\Il]+", "+1 ", s)          # +|, + |, +\|, +I
    s = re.sub(r"strength\s*\+\s*[|\\Il]+", "strength+1", s, flags=re.I)
    s = re.sub(r"\bby\s+[|I]\b", "by 1", s)

    # Stray tilde from italics, ampersand noise
    s = s.replace("~", " ")

    # Collapse whitespace
    s = re.sub(r"\s{2,}", " ", s).strip()

    # Ensure sentence ends cleanly
    s = s.rstrip(" :,;.-_|\"'")
    if s and s[-1] not in ".!?":
        s += "."

    return s


def lib_cost(card: dict) -> str:
    parts = []
    if card.get("pool_cost"):
        parts.append(f"{card['pool_cost']} pool")
    if card.get("blood_cost"):
        parts.append(f"{card['blood_cost']} blood")
    if card.get("conviction_cost"):
        parts.append(f"{card['conviction_cost']} conviction")
    return ", ".join(parts)


def build_krcg_indexes(data: list[dict]) -> tuple[dict, dict, dict]:
    vamps = [c for c in data if "Vampire" in c.get("types", []) or "Imbued" in c.get("types", [])]
    libs = [c for c in data if c not in vamps]
    crypt_by_ng: dict[tuple[str, str], dict] = {}
    crypt_by_n: dict[str, list[dict]] = {}
    for c in vamps:
        n = norm(c.get("_name", ""))
        crypt_by_ng[(n, str(c.get("group", "")))] = c
        crypt_by_n.setdefault(n, []).append(c)
    lib_by_n: dict[str, list[dict]] = {}
    for c in libs:
        lib_by_n.setdefault(norm(c.get("_name", "")), []).append(c)
        for v in c.get("name_variants", []) or []:
            lib_by_n.setdefault(norm(v), []).append(c)
    return crypt_by_ng, crypt_by_n, lib_by_n


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: source workbook not found at {XLSX}", file=sys.stderr)
        print("Restore 'data/Draft Cube.xlsx' locally before running this script.", file=sys.stderr)
        return 1
    try:
        wb = openpyxl.load_workbook(XLSX, data_only=True)
    except Exception as e:
        print(f"ERROR: failed to open {XLSX}: {e}", file=sys.stderr)
        return 1
    krcg = load_krcg()
    crypt_by_ng, crypt_by_n, lib_by_n = build_krcg_indexes(krcg)
    draft_ocr = json.loads(DRAFT_OCR.read_text(encoding="utf-8")) if DRAFT_OCR.exists() else {}
    draft_overrides: dict[str, str] = {}
    if DRAFT_OVERRIDES.exists():
        raw = json.loads(DRAFT_OVERRIDES.read_text(encoding="utf-8"))
        for k, v in raw.items():
            if not k.startswith("_") and isinstance(v, str):
                draft_overrides[k.lower()] = v

    crypt_idx = index_dir(ROOT / CRYPT_DIR)
    crypt: list[dict] = []
    missing_crypt: list[str] = []
    for row in wb["Crypt"].iter_rows(min_row=2, values_only=True):
        count, clan, name = row[0], row[1], row[2]
        if not name:
            continue
        fname = find_crypt(crypt_idx, name)
        if not fname:
            missing_crypt.append(name)
            continue

        stem = os.path.splitext(fname)[0]
        m_file = re.search(r"g(\d+)$", stem)
        file_group = m_file.group(1) if m_file else ""
        m_name = re.match(r"^(.*?)\s+[gG](\d+)$", name)
        lookup_name, lookup_group = (m_name.group(1).strip(), m_name.group(2)) if m_name else (name, file_group)

        card = crypt_by_ng.get((norm(lookup_name), lookup_group))
        if card is None:
            ms = crypt_by_n.get(norm(lookup_name), [])
            card = ms[0] if len(ms) == 1 else None

        display_name = card.get("_name") if card else lookup_name
        crypt.append({
            "name": display_name,
            "clan": clan or "",
            "count": int(count or 0),
            "img": f"{CRYPT_DIR}/{fname}",
            "capacity": card.get("capacity") if card else None,
            "group": card.get("group") if card else file_group,
            "disciplines": card.get("disciplines") or [] if card else [],
            "title": card.get("title") if card else None,
            "text": card.get("card_text") or "" if card else "",
        })

    lib_idx = {rar: index_dir(ROOT / d) for rar, d in LIB_DIRS.items()}
    library: list[dict] = []
    missing_lib: list[str] = []
    for row in wb["Library"].iter_rows(min_row=2, values_only=True):
        count, rarity, name = row[0], row[1], row[2]
        if not name:
            continue
        idx = lib_idx.get(rarity, {})
        fname = idx.get(norm(name))
        if not fname:
            missing_lib.append(f"{rarity}: {name}")
            continue
        ms = lib_by_n.get(norm(name), [])
        uniq = {c["id"]: c for c in ms}
        card = next(iter(uniq.values())) if len(uniq) == 1 else None

        img_rel = f"{LIB_DIRS[rarity]}/{fname}"
        ocr = draft_ocr.get(img_rel, {})
        override = draft_overrides.get(name.lower())
        has_draft = bool(ocr.get("has_draft")) or bool(override)
        if override:
            draft_text = override if override.upper().startswith("DRAFT") else f"DRAFT: {override}"
        else:
            draft_text = clean_draft_snippet(ocr.get("snippet") or "") if has_draft else ""
        library.append({
            "name": name,
            "rarity": rarity,
            "count": int(count or 0),
            "img": img_rel,
            "type": ", ".join(card.get("types") or []) if card else "",
            "cost": lib_cost(card) if card else "",
            "disciplines": card.get("disciplines") or [] if card else [],
            "clans": card.get("clans") or [] if card else [],
            "text": card.get("card_text") or "" if card else "",
            "draft": has_draft,
            "draft_text": draft_text,
        })

    crypt.sort(key=lambda c: c["name"].lower())
    library.sort(key=lambda c: c["name"].lower())

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump({"crypt": crypt, "library": library}, f, ensure_ascii=False, indent=2)

    print(f"Wrote {OUT.name}: {len(crypt)} crypt + {len(library)} library")
    if missing_crypt:
        print(f"Missing crypt ({len(missing_crypt)}): {missing_crypt}")
    if missing_lib:
        print(f"Missing library ({len(missing_lib)}): {missing_lib}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
